import openpyxl

wb1 = openpyxl.load_workbook("octobre.xlsx", data_only=True)
wb2 = openpyxl.load_workbook("novembre.xlsx", data_only=True)
wb3 = openpyxl.load_workbook("decembre.xlsx", data_only=True)
''' Lecture et affichage de partie de fichier excel

#print(wb.sheetnames)
#sheet = wb["Feuil1"]
#sheet = wb[wb.sheetnames[0]]

sheet = wb.active
#cell = sheet["A1"]
#cell = sheet.cell(4,3)


for row in range(2, 7):
    cell = sheet.cell(row,2)
    print("prix : ", str(cell.value)+"€")
'''
def ajouter_data(wb, d):
    sheet = wb.active
    for row in range(2, sheet.max_row):
        nom_article = sheet.cell(row, 1).value
        if not nom_article:
            break
        total_ventes = sheet.cell(row, 4).value
        if d.get(nom_article):
            d[nom_article].append(total_ventes)
        else:
            d[nom_article] = [total_ventes]

donnes = {}
ajouter_data(wb1, donnes)
ajouter_data(wb2, donnes)
ajouter_data(wb3, donnes)

#print(donnes)

wb_sortie = openpyxl.Workbook()
sheet = wb_sortie.active
sheet["A1"] = "Article"
sheet["B1"] = "Octobre"
sheet["C1"] = "Novembre"
sheet["D1"] = "Décembre"

row = 2
for i in donnes.items():
    nom_article = i[0]
    ventes = i[1]
    sheet.cell(row, 1).value = nom_article
    for j in range(0, len(ventes)):
        sheet.cell(row, 2+j).value = ventes[j]
    row += 1

chart_ref = openpyxl.chart.Reference(sheet, min_col=2, min_row=2, max_col=sheet.max_column, max_row=2)
chart_serie =   openpyxl.chart.Series(chart_ref, title="Total ventes €")
#type de graphique
chart = openpyxl.chart.BarChart()
#titre du graphique
chart.title = "Evolution du prix des pommes"
chart.append(chart_serie)
#ajout du graphique sur la feuille
sheet.add_chart(chart, "F2")
wb_sortie.save("total_vente_trimestre.xlsx")

















